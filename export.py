import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import mysql.connector
from openpyxl import load_workbook
from configparser import ConfigParser
from pathlib import Path

class ImportExportData:
    def run_setup_wizard(self):
        # Create a new Tk instance
        root = tk.Tk()

        # Call setup wizard to create the INI file
        wizard = setupwizard(root)
        print("Setup wizard completed.")

        # Run the main loop to keep the window open until it's closed
        root.mainloop()

    def __init__(self, root, config_file='config.ini'):
        # here root is frame3 from the main.py
        self.root = root
        self.config_file = config_file
        # Storing connection values
        config_path = Path(config_file)
        if not config_path.is_file():
            print(f"Config file '{config_file}' not found. Running setup wizard...")
            self.run_setup_wizard()

        # Read configuration from INI file
        config = ConfigParser()
        config.read(config_file)

        # Storing connection values
        self.db_config = {
            'host': config.get('mysql', 'host'),
            'user': config.get('mysql', 'user'),
            'password': config.get('mysql', 'password'),
            'database': config.get('mysql', 'database')
        }

        # Create connection
        self.connection = mysql.connector.connect(**self.db_config)
        self.cursor = self.connection.cursor()

    def createbutton(self,frame3):
        self.frame3=frame3

        # Creating buttons
        self.import_button = tk.Button(self.frame3, text="Import from Excel", background="#34eb98", fg="white", command=self.import_from_excel)
        self.export_button = tk.Button(self.frame3, text="Export to Excel", background="#34eb98", fg="white", command=self.export_to_excel)
        # Ensure the buttons are packed into the provided root frame
        self.import_button.pack(pady=10)
        self.export_button.pack(pady=10)

    def connect_to_database(self):
        # Check if the connection is closed or lost, and then reconnect
        if not self.connection.is_connected():
            self.connection = mysql.connector.connect(**self.db_config)
            self.cursor = self.connection.cursor()

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            try:
                # Connect to MySQL
                self.connect_to_database()

                # Load Excel file
                workbook = load_workbook(filename=file_path)
                sheet = workbook.active

                # Replace 'your_table' and 'unique_column' with appropriate values
                unique_column_index = 0  # Index of the unique column (0 for the first column, 1 for the second, and so on)

                # Iterate through rows in the sheet
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Check if the record already exists based on the unique column
                    unique_value = row[unique_column_index]
                    self.cursor.execute(f"SELECT COUNT(*) FROM your_table WHERE unique_column = %s", (unique_value,))
                    result = self.cursor.fetchone()

                    # If the record doesn't exist, insert it
                    if result[0] == 0:
                        self.cursor.execute(f"INSERT INTO your_table VALUES {tuple(row)}")

                self.connection.commit()
                self.connection.close()

                messagebox.showinfo("Import Complete", "Data imported to MySQL successfully.")

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred during import: {str(e)}")

    def export_to_excel(self):
        try:
            # Connect to MySQL
            self.connect_to_database()

            # Replace 'your_table' with the name of your MySQL table
            self.cursor.execute(f"SELECT * FROM employees")
            results = self.cursor.fetchall()

            # Create a DataFrame from the MySQL query result
            df = pd.DataFrame(results, columns=[desc[0] for desc in self.cursor.description])

            # Ask the user to choose a file path for the exported Excel file
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

            if file_path:
                # Save the DataFrame to Excel
                df.to_excel(file_path, index=False)

                messagebox.showinfo("Export Complete", f"Data exported to {file_path} successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during export: {str(e)}")

        finally:
            self.connection.close()

# Example usage for testing
if __name__ == "__main__":
    root = tk.Tk()
    root.geometry('800x600')
    root.title('Data Import/Export')
    
    # Create an object of ImportExportData
    data_import_export = ImportExportData(root)
    
    root.mainloop()
